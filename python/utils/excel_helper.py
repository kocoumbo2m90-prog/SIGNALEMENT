import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelHelper:
    @staticmethod
    def get_excel_path():
        return os.path.join(os.getcwd(), "registre_signalements.xlsx")

    @classmethod
    def init_excel(cls):
        path = cls.get_excel_path()
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Signalements"

        # Activer le quadrillage visible
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "ID", "Date", "Titre", "Description", "Catégorie", "Gravité", "Statut",
            "Latitude", "Longitude", "Adresse", "Audio (Lien)", "Médias/Photos (Lien)", 
            "Anonyme", "Nom déclarant", "Email", "Téléphone", "Notes Admin"
        ]

        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        ws.row_dimensions[1].height = 28

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

            # Initialisation d'une largeur par défaut pour éviter le recalcul total plus tard
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(len(header) + 4, 12)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        wb.save(path)

    @classmethod
    def update_or_add_report(cls, report):
        """Ajoute ou met à jour un signalement dans le fichier Excel"""
        cls.init_excel()
        path = cls.get_excel_path()
        wb = load_workbook(path)
        ws = wb.active

        audio_link = os.path.basename(report.audio_uri) if report.audio_uri else "Aucun"
        media_link = os.path.basename(report.media_uri) if report.media_uri else "Aucun"
        
        row_data = [
            report.id,
            report.timestamp.strftime("%Y-%m-%d %H:%M:%S") if report.timestamp else "",
            report.title,
            report.description,
            report.category,
            report.severity,
            report.status or "Nouveau",
            report.latitude,
            report.longitude,
            report.location_address,
            audio_link,
            media_link,
            "Oui" if report.is_anonymous else "Non",
            report.reporter_name if not report.is_anonymous else "N/A",
            report.reporter_email if not report.is_anonymous else "N/A",
            report.reporter_phone if not report.is_anonymous else "N/A",
            report.admin_notes or ""
        ]

        # Recherche de l'ID existant
        target_row = None
        # OPTIMISATION : Si le tableau est TRÈS grand, bouclez plutôt sur ws['A'] 
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == report.id:
                target_row = row
                break

        if target_row is None:
            target_row = ws.max_row + 1

        # Définition unique de la bordure pour éviter de recréer l'objet en boucle
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        standard_font = Font(name="Segoe UI", size=10)

        # Écriture des données et ajustement ciblé des colonnes
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            
            # Alignement
            if col_idx in [1, 2, 5, 6, 7, 8, 9, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            cell.font = standard_font
            cell.border = thin_border

            # OPTIMISATION IA : Ajustement de largeur intelligent (uniquement pour la cellule modifiée)
            col_letter = get_column_letter(col_idx)
            current_width = ws.column_dimensions[col_letter].width or 12
            new_len = len(str(value or '')) + 3
            
            # On n'agrandit la colonne que si la nouvelle valeur dépasse la taille actuelle
            if new_len > current_width:
                ws.column_dimensions[col_letter].width = max(new_len, 12)

        wb.save(path)